'''
Created on 1 Dec 2016

@author: Michael
'''

import os,sys,json,csv
import hashlib
from openpyxl import Workbook,load_workbook
from scan import scan_files, parse_args, log

def md5(fname):
    hash_md5 = hashlib.md5()
    with open(fname, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

data = {}
xlsx = {}
#columns = ['1_2','1_3','1_4','2_7','2_16','2_22','3_8','3_32','4_10','4_32']
columns = ['speaker']

def pre_dir(**kw):
    log("Scanning: "+kw['dirpath'],kw['outfile'])

def on_file(**kw):
    ''' Searches for particular file names and counts to provide data on how things look '''
    global data
    global columns
    try:
        file,ext = kw['file'].split('.')
    except:
        #odd file, ignore
        log("Odd File: "+os.path.join(kw['dirpath'],kw['file']),kw['outfile'])
        return
    if ext=='xlsx':
        #We have an excel file. We should collect the data.
        try:
            comps = file.split('_')
            speaker = comps[1]+'_'+comps[2]
            if not comps[0]=='MAUS':
                #Not interesting excel file, skip
                return
            
        except:
            #Not an interesting excel file, skip
            return
        
        if not speaker in xlsx:
            xlsx[speaker] = {}
        try:
            wb = load_workbook(os.path.join(kw['dirpath'],kw['file']))
        except:
            #odd file, ignore
            log("Couldn't load Excel File: "+os.path.join(kw['dirpath'],kw['file']),kw['outfile'])
            return
        ws = wb.active
        
        #They all use rows 2 to 60 (inclusive)
        for row in range(2,61):
            #Get numbers from Col A as keys
            key = ws['A%d'%row].value
            #If something in Col C then value is true, false otherwise
            val = 'false'
            if len(str(ws['C%d'%row].value))>0:
                val = 'true'
            
            xlsx[speaker][key] = val
            
        log('Processed Excel File: '+kw['file'],kw['outfile'])
        
    else:
        try:
            comps = file.split('_')
            speaker = comps[0]+'_'+comps[1]
            session = comps[2]+'_'+comps[3]
            item = comps[4][:3] #This may have -ch6-speaker or -n-n-n behind it
        except:
            #Not the file we were looking for
            #move along, move along
            return
        
        #Initialize anything that needs it
        if not speaker in data:
            data[speaker] = {}
        
        if not session in data[speaker]:
            data[speaker][session]={}
        
        #For duplicate files
        if not item in data[speaker][session]:
            data[speaker][session][item]={}
        
        #Add in the file data
        path = os.path.join(kw['dirpath'],kw['file'])
        hash = md5(path)
        
        
        if not hash in data[speaker][session][item]:
            data[speaker][session][item][hash] = []
        
        data[speaker][session][item][hash].append(path)
        
        col_name = "%s %s" % (session,ext)
        
        if not col_name in columns:
            columns.append(col_name)

def on_finish(**kw):
    ''' output all data '''
    global data
    
    log("Finished Scanning Files",kw['outfile'])
    
    #generate output
    with open(kw['output_file']+'.json','w') as file:
        file.write(json.dumps(data))
        
    with open(kw['output_file']+'_excel.json','w') as file:
        file.write(json.dumps(xlsx))
        
    log("Finished Generating json output",kw['outfile'])
    
    csv_rows = []
    
    for speaker in data:
        tmp = {'speaker':speaker}
        for session in data[speaker]:
            for item in data[speaker][session]:
                ext_list = []
                for hash in data[speaker][session][item]:
                    ext = data[speaker][session][item][hash][0].split('.')[-1]
                    #prevent updated or older files for the same speaker/session form being counted
                    if not ext in ext_list:
                        ext_list.append(ext)
                        col_name = "%s %s" % (session,ext)
                        if not col_name in tmp:
                            tmp[col_name] = 0
                        tmp[col_name] += 1
        csv_rows.append(tmp)
    
    
    
    with open(kw['output_file']+'.csv','w') as file:
        dict_writer = csv.DictWriter(file,lineterminator='\n',fieldnames=columns)
        dict_writer.writeheader()
        dict_writer.writerows(csv_rows)
    
    log("Finished Generating csv output",kw['outfile'])

if __name__ == '__main__':
    vars = parse_args(extra_usage="""Scans provided directories and refactors them into a new directory structure.""")
    output_file='annotation_data'
    with open(output_file+'.log','w') as file:
        print "Starting Scan"
        scan_files(vars['root'], outfile=file, function=on_file,on_finish=on_finish,pre_dir=pre_dir,output_file=output_file)
        